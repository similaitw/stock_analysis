"""
Excel 報告生成器
Excel Report Generator
"""
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelReportGenerator:
    """Excel 報告生成器"""
    
    def __init__(self, output_dir: str = "reports/output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_stock_analysis_report(
        self,
        ticker: str,
        df: pd.DataFrame,
        current_price: float = None,
        indicators: dict = None
    ) -> str:
        """
        生成個股分析報告
        
        Args:
            ticker: 股票代碼
            df: 股價資料
            current_price: 當前價格
            indicators: 技術指標資料
        
        Returns:
            報告檔案路徑
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "個股分析"
        
        # 標題
        ws['A1'] = f"{ticker} 個股分析報告"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 基本資訊
        row = 4
        if current_price:
            ws[f'A{row}'] = "當前價格"
            ws[f'B{row}'] = current_price
            row += 1
        
        # 股價資料
        row += 2
        ws[f'A{row}'] = "歷史股價資料"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # 寫入 DataFrame
        for r_idx, row_data in enumerate(dataframe_to_rows(df, index=True, header=True), row):
            for c_idx, value in enumerate(row_data, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 儲存檔案
        filename = f"{ticker}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        return str(filepath)
    
    def generate_backtest_report(
        self,
        ticker: str,
        strategy: str,
        results: dict,
        trades: pd.DataFrame = None
    ) -> str:
        """
        生成回測報告
        
        Args:
            ticker: 股票代碼
            strategy: 策略名稱
            results: 回測結果
            trades: 交易明細
        
        Returns:
            報告檔案路徑
        """
        wb = Workbook()
        
        # 摘要頁
        ws_summary = wb.active
        ws_summary.title = "回測摘要"
        
        ws_summary['A1'] = f"{ticker} - {strategy} 回測報告"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 績效指標
        row = 4
        ws_summary[f'A{row}'] = "績效指標"
        ws_summary[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        for key, value in results.items():
            ws_summary[f'A{row}'] = key
            ws_summary[f'B{row}'] = value
            row += 1
        
        # 交易明細頁
        if trades is not None and not trades.empty:
            ws_trades = wb.create_sheet("交易明細")
            
            for r_idx, row_data in enumerate(dataframe_to_rows(trades, index=False, header=True), 1):
                for c_idx, value in enumerate(row_data, 1):
                    ws_trades.cell(row=r_idx, column=c_idx, value=value)
        
        # 儲存檔案
        filename = f"{ticker}_{strategy}_backtest_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        return str(filepath)
    
    def generate_scanner_report(
        self,
        strategy: str,
        market: str,
        results: pd.DataFrame
    ) -> str:
        """
        生成選股報告
        
        Args:
            strategy: 策略名稱
            market: 市場範圍
            results: 選股結果
        
        Returns:
            報告檔案路徑
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "選股結果"
        
        # 標題
        ws['A1'] = f"{strategy} 選股報告"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"市場範圍: {market}"
        ws['A3'] = f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"符合條件: {len(results)} 檔"
        
        # 選股結果
        row = 6
        ws[f'A{row}'] = "選股結果"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # 寫入 DataFrame
        for r_idx, row_data in enumerate(dataframe_to_rows(results, index=False, header=True), row):
            for c_idx, value in enumerate(row_data, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 儲存檔案
        filename = f"{strategy}_{market}_scanner_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        return str(filepath)
